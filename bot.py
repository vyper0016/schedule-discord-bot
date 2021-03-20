
# Added support for classmates that have the same schedule
# Added Dictionary entry for subjects to skip - 4/12/2020

import json
import openpyxl
import datetime
from os import remove
import discord
from discord.ext import commands
from discord.ext.commands import CommandNotFound

client = commands.Bot(command_prefix='!')

with open('users.json') as f:
    users = json.load(f)
    
with open('calendar.json', 'r') as cf:
    calendar = json.load(cf)


def list_to_string(lst: list):
    string = ''
    for i, c in zip(lst, range(len(lst) + 1)):
        if c == len(lst) - 1:
            string += 'and ' + i
            return string
        string += i + ', '


def next_week(week):

    n = week[:-1]
    n = str(int(n) + 1)
    if week[-1] == 'B':
        k = 'A'
    else:
        k = 'B'

    return n + k


def get_week(day):

    x = calendar[day]['week']

    if calendar[day]['day'] != 'Sunday':
        return x

    return next_week(x)


def day_in_weeks(day, n):

    week = get_week(day)
    for i in range(n):
        week = next_week(week)

    for i in calendar:
        if calendar[i]['week'] == week:
            return i

    return 'Week not found'


def days(day):
    lst = []
    w = get_week(day)
    for i in calendar:
        if len(lst) == 6:
            return lst

        if calendar[i]['week'] == w:
            lst.append(i)


def th_number(x):
    x = int(x)
    if x < 21:
        if x == 1:
            return str(x) + 'st'
        elif x == 2:
            return str(x) + 'nd'
        elif x == 3:
            return str(x) + 'rd'
    else:
        u = x % 10
        if u == 1:
            return str(x) + 'st'
        elif u == 2:
            return str(x) + 'nd'
        elif u == 3:
            return str(x) + 'rd'
    return str(x) + 'th'


def title(day):
    days_lst = days(day)
    w = get_week(day)
    d1 = days_lst[0].split('/')[1]
    d2 = days_lst[5].split('/')[1]
    wl = w[-1]
    wn = w[:-1]
    month1 = datetime.date(1900, int(days_lst[0].split('/')[0]), 1).strftime('%b')
    month2 = datetime.date(1900, int(days_lst[0].split('/')[0]), 1).strftime('%b')

    return '{0} week ({1} {2} - {3} {4}) ({5})'.format(
        th_number(wn), th_number(d1), month1, th_number(d2), month2, wl)


def next_column(c):
    le = chr(ord(c[0]) - 1) + c[1]
    nl = chr(ord(c[0]) + 1) + c[1]
    return c + ':' + nl, le + ':' + c


def write_xlsx(day, idi):

    cells = users[idi]['cells']

    with open(cells, 'r') as cef:
        cells = json.load(cef)

    sample = users[idi]['sample']

    wb = openpyxl.load_workbook(sample)
    ws = wb['Sheet1']

    with open('data' + str(users[idi]['group']) + '.json', 'r') as df:
        data = json.load(df)

    days_list = days(day)

#   d : day

    for d in days_list:
        if not data[d]['set']:

            for i in cells[calendar[d]['day']]:

                if not ws[i].value:
                    continue

                if strip_subject(ws[i].value) in users[idi]['toSkip']:
                    # -x + 3  To flip from 1 to 2 and vice-versa
                    # checks if other group studies before ignoring (skipping) subject
                    with open('data' + str(- users[idi]['group'] + 3) + '.json', 'r') as f2:
                        data2 = json.load(f2)

                    if data2[d]['set']:
                        continue

                ws[i] = ''

    file_name = title(day)
    
    ws[cells['title']] = file_name

    file_name += ' Schedule.xlsx'
    wb.save(filename=file_name)
    return file_name


def convert_day(day: str):
    
    # Includes wrong format detection.
    
    if day.count('/') != 2:
        return 'invalid Date Format'

    lst = day.split('/')
    output_day = ''

    for i in lst:
        try:
            int(i)
        except ValueError:
            return 'invalid Date Format'

    if int(lst[1]) > 12:
        return 'invalid Month Format'
    elif len(lst[1]) == 2:
        if lst[1][:1] == '0':
            output_day += lst[1][1:] + '/'
        else:
            output_day += lst[1] + '/'
    elif len(lst[1]) == 1:
        output_day += lst[1] + '/'
    else:
        return 'invalid Month Format'

    if int(lst[0]) > 31:
        return 'invalid Day Format'
    elif len(lst[0]) == 2:
        if lst[0][:1] == '0':
            output_day += lst[0][1:] + '/'
        else:
            output_day += lst[0] + '/'
    elif len(lst[0]) == 1:
        output_day += lst[0] + '/'
    else:
        return 'invalid Day Format'

    if len(lst[2]) == 4:
        output_day += lst[2][2:]
    elif len(lst[2]) == 2:
        output_day += lst[2]
    else:
        return 'invalid Year Format'

    return output_day


def classmate(uid):
    uid = str(uid)

    if uid in users:
        return uid

    for i in users:
        if uid in list(map(str, users[i]['classmates'])):
            return i


def gather_subjects(uid):
    sample = users[str(uid)]['sample']
    with open(users[str(uid)]['cells'], 'r') as cef:
        cells = json.load(cef)

    wb = openpyxl.load_workbook(sample)
    ws = wb['Sheet1']

    d = None

    for d in calendar:
        if calendar[d]['day'] == 'Monday':
            break

    day_list = days(d)
    subjects = []
    for i in day_list:
        for c in cells[calendar[i]['day']]:
            subject = ws[c].value
            if subject in subjects:
                continue
            subjects.append(subject)

    subjects = list(map(strip_subject, subjects))

    return subjects


def strip_subject(s: str):
    ns = ''
    s = s.strip()
    for i in s:
        sc = i.casefold()
        if not (ord('a') <= ord(sc) <= ord('z')):
            continue
        ns += sc
    return ns


@client.event
async def on_ready():
    print('Started')


@client.event
async def on_command_error(_, error):
    if isinstance(error, CommandNotFound):
        return
    raise error


@client.command(aliases=['sch', 's'])
async def schedule(ctx, input_day=None):

    uid = classmate(ctx.author.id)

    if not uid:
        await ctx.send('Ayy bro you not in the database!')
        return

    today = datetime.datetime.now().strftime("%d/%m/%Y")
    input_day = input_day or today

    try:
        nweek = int(input_day)
    except ValueError:
        nweek = None

    if nweek:

        if nweek < 0:
            await ctx.send('You trinna give me negative week numbers? :thinking:\nNo one cares about the past '
                           'bruv\nFormat should be: `<number of weeks>`')
            return

        converted_day = day_in_weeks(convert_day(today), nweek)

        if converted_day == 'Week not found':
            await ctx.send('Could not find week :thinking:')
            return

    else:
        converted_day = convert_day(input_day)

    if '/' not in converted_day:
        await ctx.send('"{}"? That\'s an {} :thinking:\nFormat should be: `DD/MM/YYYY`\nOr you can specify the number'
                       ' of weeks to skip'.format(input_day, converted_day))
        return
    if converted_day not in calendar:
        print(converted_day)
        await ctx.send(f'{input_day} was not found :thinking:')
        return

    n = write_xlsx(converted_day, uid)

    embed = discord.Embed()

    embed.set_footer(text=f"Goodluck out there, {users[uid]['name']}")

    if nweek:
        embed.title = f'Generated week\'s schedule for: {n[:-14]}'

    else:
        dn = calendar[converted_day]['day']
        embed.title = "Generated week's schedule for: {} ({})".format(input_day, dn)
        if dn == 'Sunday':
            embed.set_footer(text="Date was a Sunday, showing schedule for next week")

    with open(n, 'rb') as ff:
        file = discord.File(ff, filename=n)
        await ctx.send(file=file, embed=embed)

    remove(n)


@client.command(aliases=['wsch', 'w', 'wholeSchedule', 'whole', 'wh', 'who'])
async def whole_schedule(ctx):
    uid = classmate(ctx.author.id)

    if not uid:
        await ctx.send('Ayy bro you not in the database!')
        return

    image = users[str(uid)]['schedulePicture']

    file = discord.File(image, filename="image.png")
    embed = discord.Embed()
    embed.set_author(name='Here\'s the whole schedule:')
    embed.set_image(url="attachment://image.png")
    await ctx.send(file=file, embed=embed)

    n = users[str(ctx.author.id)]['sample']

    with open(n, 'rb') as fx:
        file = discord.File(fx, filename=n)
        await ctx.send(file=file)


@client.command(aliases=['sk', 'skip', 'ts'])
async def to_skip(ctx, subject_input: str):

    uid = classmate(ctx.author.id)

    if not uid:
        await ctx.send('Ayy bro you not in the database!')
        return

    subjects = gather_subjects(uid)
    subject = strip_subject(subject_input)

    if subject not in subjects:
        await ctx.send(f'I don\'t think {subject_input} is a valid subject :thinking:\n here\'s a list of your '
                       f'subjects:\n{list_to_string(subjects)}')
        return 
    
    if subject not in users[uid]['toSkip']:
        users[uid]['toSkip'].append(subject)
        with open('users.json', 'w') as uf:
            json.dump(users, uf, indent=3)

        await ctx.send(f'Added {subject_input} to your skipping list!'
                       '\n To remove it simply type the same command again')
        return

    users[uid]['toSkip'].remove(subject)
    with open('users.json', 'w') as uf:
        json.dump(users, uf, indent=3)

    await ctx.send(f'removed {subject_input} from your skipping list.')


@client.command()
async def ping(ctx):
    await ctx.send(f'Pong :ping_pong:{round(client.latency * 1000)}ms')


@to_skip.error
async def switch_error(ctx, error):
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send('you have to include the subject! :thinking:')
        return

client.run('')
